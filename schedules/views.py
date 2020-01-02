# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render
from django.views.generic import (TemplateView, CreateView, ListView)
from rest_framework import viewsets, generics


from . import serializers, models, forms
from django.db.models import Q
import openpyxl
from datetime import datetime, timedelta

def JadwalKuliah(request):
    return render(request, 'jadwalkuliah.html', {})

def Kuliah(request):
    if "POST" == request.method:
        query1 = request.POST.get('key1')
        query2 = request.POST.get('key2')
        query3 = request.POST.get('key3')
        query4 = request.POST.get('key4')
        data1 = models.Kuliah.objects.filter(
            Q(day__contains=query1) |
            Q(full_name__contains=query1) |
            Q(initial__contains=query1) |
            Q(subject_code__contains=query1) |
            Q(subject_name__contains=query1) |
            Q(cohort_name__contains=query1) |
            Q(room_no__contains=query1)
        )
        data2 = models.Kuliah.objects.filter(
            Q(day__contains=query2) |
            Q(full_name__contains=query2) |
            Q(initial__contains=query2) |
            Q(subject_code__contains=query2) |
            Q(subject_name__contains=query2) |
            Q(cohort_name__contains=query2) |
            Q(room_no__contains=query2)
        )
        data3 = models.Kuliah.objects.filter(
            Q(day__contains=query3) |
            Q(full_name__contains=query3) |
            Q(initial__contains=query3) |
            Q(subject_code__contains=query3) |
            Q(subject_name__contains=query3) |
            Q(cohort_name__contains=query3) |
            Q(room_no__contains=query3)
        )
        data4 = models.Kuliah.objects.filter(
            Q(day__contains=query4) |
            Q(full_name__contains=query4) |
            Q(initial__contains=query4) |
            Q(subject_code__contains=query4) |
            Q(subject_name__contains=query4) |
            Q(cohort_name__contains=query4) |
            Q(room_no__contains=query4)
        )
        if (query2 == "" and query3 == "" and query4 == "" ):
            data = data1
        elif(query3 == "" and query4 == ""):
            data = data1 | data2
        elif(query4 == ""):
            data = data1 | data2 | data3
        else:
            data = data1 | data2 | data3 | data4

    else:
        data = models.Kuliah.objects.all()

    # now = datetime.now()
    # a = now.strftime("%w")
    # pengurang = int(a)-1
    # tanggal_mulai = now+timedelta(days=-pengurang)

    kuliah_senin = data.filter(day__contains="Monday").order_by("class_date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_senin:
        if iterate == 0:
            week1_tgl = item.class_date
            iterate = iterate +1
        else:
            if(item.class_date != week1_tgl):
                week2_tgl = item.class_date
                break;

    tanggal_senin= week1_tgl
    tanggal_senin2 = week2_tgl
    senin_1 = kuliah_senin.filter(class_date__date=tanggal_senin)
    senin_2 = kuliah_senin.filter(class_date__date=tanggal_senin2)
    kuliah_senin_slot1 = senin_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_senin_slot2 = senin_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_senin_slot3 = senin_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_senin_slot4 = senin_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_senin_slot5 = senin_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_senin2_slot1 = senin_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_senin2_slot2 = senin_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_senin2_slot3 = senin_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_senin2_slot4 = senin_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_senin2_slot5 = senin_2.filter(start_time__range=("17:50", "19:30"))

    kuliah_selasa = data.filter(day__contains="Tuesday").order_by("class_date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_selasa:
        if iterate == 0:
            week1_tgl = item.class_date
            iterate = iterate +1
        else:
            if(item.class_date != week1_tgl):
                week2_tgl = item.class_date
                break;
    tanggal_selasa =  week1_tgl
    tanggal_selasa2 = week2_tgl

    selasa_1 = kuliah_selasa.filter(class_date__date=tanggal_selasa)
    selasa_2 = kuliah_selasa.filter(class_date__date=tanggal_selasa2)
    kuliah_selasa_slot1 = selasa_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_selasa_slot2 = selasa_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_selasa_slot3 = selasa_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_selasa_slot4 = selasa_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_selasa_slot5 = selasa_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_selasa2_slot1 = selasa_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_selasa2_slot2 = selasa_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_selasa2_slot3 = selasa_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_selasa2_slot4 = selasa_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_selasa2_slot5 = selasa_2.filter(start_time__range=("17:50", "19:30"))

    kuliah_rabu = data.filter(day__contains="Wednesday").order_by("class_date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_rabu:
        if iterate == 0:
            week1_tgl = item.class_date
            iterate = iterate +1
        else:
            if(item.class_date != week1_tgl):
                # print("ee")
                week2_tgl = item.class_date
                break;
    tanggal_rabu =  week1_tgl
    tanggal_rabu2= week2_tgl
    rabu_1 = kuliah_rabu.filter(class_date__date=tanggal_rabu)
    rabu_2 = kuliah_rabu.filter(class_date__date=tanggal_rabu2)
    kuliah_rabu_slot1 = rabu_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_rabu_slot2 = rabu_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_rabu_slot3 = rabu_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_rabu_slot4 = rabu_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_rabu_slot5 = rabu_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_rabu2_slot1 = rabu_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_rabu2_slot2 = rabu_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_rabu2_slot3 = rabu_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_rabu2_slot4 = rabu_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_rabu2_slot5 = rabu_2.filter(start_time__range=("17:50", "19:30"))

    kuliah_kamis = data.filter(day__contains="Thursday").order_by("class_date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_kamis:
        if iterate == 0:
            week1_tgl = item.class_date
            iterate = iterate +1
        else:
            if(item.class_date != week1_tgl):
                # print("ee")
                week2_tgl = item.class_date
                break;
    tanggal_kamis =  week1_tgl
    tanggal_kamis2= week2_tgl
    kamis_1 = kuliah_kamis.filter(class_date__date=tanggal_kamis)
    kamis_2 = kuliah_kamis.filter(class_date__date=tanggal_kamis2)
    kuliah_kamis_slot1 = kamis_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_kamis_slot2 = kamis_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_kamis_slot3 = kamis_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_kamis_slot4 = kamis_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_kamis_slot5 = kamis_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_kamis2_slot1 = kamis_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_kamis2_slot2 = kamis_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_kamis2_slot3 = kamis_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_kamis2_slot4 = kamis_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_kamis2_slot5 = kamis_2.filter(start_time__range=("17:50", "19:30"))

    kuliah_jumat = data.filter(day__contains="Friday").order_by("class_date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_jumat:
        if iterate == 0:
            week1_tgl = item.class_date
            iterate = iterate +1
        else:
            if(item.class_date != week1_tgl):
                # print("horeeee")
                week2_tgl = item.class_date
                break;
    tanggal_jumat =  week1_tgl
    tanggal_jumat2= week2_tgl
    jumat_1 = kuliah_jumat.filter(class_date__date=tanggal_jumat)
    jumat_2 = kuliah_jumat.filter(class_date__date=tanggal_jumat2)
    kuliah_jumat_slot1 = jumat_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_jumat_slot2 = jumat_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_jumat_slot3 = jumat_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_jumat_slot4 = jumat_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_jumat_slot5 = jumat_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_jumat2_slot1 = jumat_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_jumat2_slot2 = jumat_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_jumat2_slot3 = jumat_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_jumat2_slot4 = jumat_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_jumat2_slot5 = jumat_2.filter(start_time__range=("17:50", "19:30"))

    #check empty slot for Senin
    if (tanggal_senin != None) and (tanggal_senin2 ==  None) :
        flag = False
        if tanggal_selasa != None:
            if tanggal_selasa < tanggal_senin:
                flag = True
        elif tanggal_rabu != None:
            if tanggal_rabu < tanggal_senin:
                flag = True
        elif tanggal_kamis != None:
            if tanggal_kamis < tanggal_senin:
                flag = True
        elif tanggal_jumat != None:
            if tanggal_jumat < tanggal_senin:
                flag = True

        if flag == True :
            tanggal_senin2 = tanggal_senin
            tanggal_senin = None
            kuliah_senin2_slot1 = kuliah_senin_slot1
            kuliah_senin2_slot2 = kuliah_senin_slot2
            kuliah_senin2_slot3 = kuliah_senin_slot3
            kuliah_senin2_slot4 = kuliah_senin_slot4
            kuliah_senin2_slot5 = kuliah_senin_slot5
            kuliah_senin_slot1 = []
            kuliah_senin_slot2 = []
            kuliah_senin_slot3 = []
            kuliah_senin_slot4 = []
            kuliah_senin_slot5 = []

    if (tanggal_selasa != None) and (tanggal_selasa2 ==  None) :
        flag = False
        if tanggal_rabu != None:
            if tanggal_rabu < tanggal_selasa:
                tanggal_selasa2 =  tanggal_selasa
        elif tanggal_kamis != None:
            if tanggal_kamis < tanggal_selasa:
                flag = True
        elif tanggal_jumat != None:
            if tanggal_jumat < tanggal_selasa:
                tanggal_selasa2 =  tanggal_selasa

        if flag == True :
            tanggal_selasa2 = tanggal_selasa
            tanggal_selasa = None
            kuliah_selasa2_slot1 = kuliah_selasa_slot1
            kuliah_selasa2_slot2 = kuliah_selasa_slot2
            kuliah_selasa2_slot3 = kuliah_selasa_slot3
            kuliah_selasa2_slot4 = kuliah_selasa_slot4
            kuliah_selasa2_slot5 = kuliah_selasa_slot5
            kuliah_selasa_slot1 = []
            kuliah_selasa_slot2 = []
            kuliah_selasa_slot3 = []
            kuliah_selasa_slot4 = []
            kuliah_selasa_slot5 = []

    if (tanggal_rabu != None) and (tanggal_rabu2 ==  None) :
        flag = False
        if tanggal_kamis != None:
            if tanggal_kamis < tanggal_rabu:
                flag = True
        elif tanggal_jumat != None:
            if tanggal_jumat < tanggal_rabu:
                flag = True

        if flag == True :
            tanggal_rabu2 = tanggal_rabu
            tanggal_rabu = None
            kuliah_rabu2_slot1 = kuliah_rabu_slot1
            kuliah_rabu2_slot2 = kuliah_rabu_slot2
            kuliah_rabu2_slot3 = kuliah_rabu_slot3
            kuliah_rabu2_slot4 = kuliah_rabu_slot4
            kuliah_rabu2_slot5 = kuliah_rabu_slot5
            kuliah_rabu_slot1 = []
            kuliah_rabu_slot2 = []
            kuliah_rabu_slot3 = []
            kuliah_rabu_slot4 = []
            kuliah_rabu_slot5 = []

    if (tanggal_kamis != None) and (tanggal_kamis2 ==  None) :
        flag = False
        if tanggal_jumat != None:
            if tanggal_jumat < tanggal_kamis:
                flag = True

        if flag == True :
            tanggal_kamis2 = tanggal_rabu
            tanggal_kamis = None
            kuliah_kamis2_slot1 = kuliah_kamis_slot1
            kuliah_kamis2_slot2 = kuliah_kamis_slot2
            kuliah_kamis2_slot3 = kuliah_kamis_slot3
            kuliah_kamis2_slot4 = kuliah_kamis_slot4
            kuliah_kamis2_slot5 = kuliah_kamis_slot5
            kuliah_kamis_slot1 = []
            kuliah_kamis_slot2 = []
            kuliah_kamis_slot3 = []
            kuliah_kamis_slot4 = []
            kuliah_kamis_slot5 = []

    if (tanggal_jumat != None) and (tanggal_jumat2 ==  None) :
        flag = False
        if tanggal_senin2!= None:
            if tanggal_senin2 < tanggal_jumat:
                flag = True
        elif tanggal_selasa2 != None:
            if tanggal_selasa2 < tanggal_jumat:
                flag = True
        elif tanggal_rabu2 != None:
            if tanggal_rabu2 < tanggal_jumat:
                flag = True
        elif tanggal_kamis2 != None:
            if tanggal_kamis2 < tanggal_jumat:
                flag = True

        if flag == True :
            tanggal_jumat2 = tanggal_jumat
            tanggal_jumat = None
            kuliah_jumat2_slot1 = kuliah_jumat_slot1
            kuliah_jumat2_slot2 = kuliah_jumat_slot2
            kuliah_jumat2_slot3 = kuliah_jumat_slot3
            kuliah_jumat2_slot4 = kuliah_jumat_slot4
            kuliah_jumat2_slot5 = kuliah_jumat_slot5
            kuliah_jumat_slot1 = []
            kuliah_jumat_slot2 = []
            kuliah_jumat_slot3 = []
            kuliah_jumat_slot4 = []
            kuliah_jumat_slot5 = []



    return render(request, 'kuliah.html', {
                "tanggal_senin":tanggal_senin,
                "tanggal_selasa":tanggal_selasa,
                "tanggal_rabu":tanggal_rabu,
                "tanggal_kamis":tanggal_kamis,
                "tanggal_jumat":tanggal_jumat,
                "tanggal_senin2":tanggal_senin2,
                "tanggal_selasa2":tanggal_selasa2,
                "tanggal_rabu2":tanggal_rabu2,
                "tanggal_kamis2":tanggal_kamis2,
                "tanggal_jumat2":tanggal_jumat2,

                "kuliah_senin_slot1":kuliah_senin_slot1,
                "kuliah_senin_slot2":kuliah_senin_slot2,
                "kuliah_senin_slot3":kuliah_senin_slot3,
                "kuliah_senin_slot4":kuliah_senin_slot4,
                "kuliah_senin_slot5":kuliah_senin_slot5,
                "kuliah_senin2_slot1":kuliah_senin2_slot1,
                "kuliah_senin2_slot2":kuliah_senin2_slot2,
                "kuliah_senin2_slot3":kuliah_senin2_slot3,
                "kuliah_senin2_slot4":kuliah_senin2_slot4,
                "kuliah_senin2_slot5":kuliah_senin2_slot5,

                "kuliah_selasa_slot1":kuliah_selasa_slot1,
                "kuliah_selasa_slot2":kuliah_selasa_slot2,
                "kuliah_selasa_slot3":kuliah_selasa_slot3,
                "kuliah_selasa_slot4":kuliah_selasa_slot4,
                "kuliah_selasa_slot5":kuliah_selasa_slot5,
                "kuliah_selasa2_slot1":kuliah_selasa2_slot1,
                "kuliah_selasa2_slot2":kuliah_selasa2_slot2,
                "kuliah_selasa2_slot3":kuliah_selasa2_slot3,
                "kuliah_selasa2_slot4":kuliah_selasa2_slot4,
                "kuliah_selasa2_slot5":kuliah_selasa2_slot5,

                "kuliah_rabu_slot1":kuliah_rabu_slot1,
                "kuliah_rabu_slot2":kuliah_rabu_slot2,
                "kuliah_rabu_slot3":kuliah_rabu_slot3,
                "kuliah_rabu_slot4":kuliah_rabu_slot4,
                "kuliah_rabu_slot5":kuliah_rabu_slot5,
                "kuliah_rabu2_slot1":kuliah_rabu2_slot1,
                "kuliah_rabu2_slot2":kuliah_rabu2_slot2,
                "kuliah_rabu2_slot3":kuliah_rabu2_slot3,
                "kuliah_rabu2_slot4":kuliah_rabu2_slot4,
                "kuliah_rabu2_slot5":kuliah_rabu2_slot5,

                "kuliah_kamis_slot1":kuliah_kamis_slot1,
                "kuliah_kamis_slot2":kuliah_kamis_slot2,
                "kuliah_kamis_slot3":kuliah_kamis_slot3,
                "kuliah_kamis_slot4":kuliah_kamis_slot4,
                "kuliah_kamis_slot5":kuliah_kamis_slot5,
                "kuliah_kamis2_slot1":kuliah_kamis2_slot1,
                "kuliah_kamis2_slot2":kuliah_kamis2_slot2,
                "kuliah_kamis2_slot3":kuliah_kamis2_slot3,
                "kuliah_kamis2_slot4":kuliah_kamis2_slot4,
                "kuliah_kamis2_slot5":kuliah_kamis2_slot5,

                "kuliah_jumat_slot1":kuliah_jumat_slot1,
                "kuliah_jumat_slot2":kuliah_jumat_slot2,
                "kuliah_jumat_slot3":kuliah_jumat_slot3,
                "kuliah_jumat_slot4":kuliah_jumat_slot4,
                "kuliah_jumat_slot5":kuliah_jumat_slot5,
                "kuliah_jumat2_slot1":kuliah_jumat2_slot1,
                "kuliah_jumat2_slot2":kuliah_jumat2_slot2,
                "kuliah_jumat2_slot3":kuliah_jumat2_slot3,
                "kuliah_jumat2_slot4":kuliah_jumat2_slot4,
                "kuliah_jumat2_slot5":kuliah_jumat2_slot5,
        }
    )

def NewKuliah(request):
    if "POST" == request.method:
        query1 = request.POST.get('key1')
        query2 = request.POST.get('key2')
        query3 = request.POST.get('key3')
        query4 = request.POST.get('key4')
        data1 = models.NewKuliahRevise.objects.filter(
            Q(day__contains=query1) |
            Q(faculty_name__contains=query1) |
            Q(nik__contains=query1) |
            Q(subject_name__contains=query1) |
            Q(event_name__contains=query1) |
            Q(section_name__contains=query1) |
            Q(room__contains=query1)
        )
        data2 = models.NewKuliahRevise.objects.filter(
            Q(day__contains=query2) |
            Q(faculty_name__contains=query2) |
            Q(nik__contains=query2) |
            Q(subject_name__contains=query2) |
            Q(event_name__contains=query2) |
            Q(section_name__contains=query2) |
            Q(room__contains=query2)
        )
        data3 = models.NewKuliahRevise.objects.filter(
            Q(day__contains=query3) |
            Q(faculty_name__contains=query3) |
            Q(nik__contains=query3) |
            Q(subject_name__contains=query3) |
            Q(event_name__contains=query3) |
            Q(section_name__contains=query3) |
            Q(room__contains=query3)
        )
        data4 = models.NewKuliahRevise.objects.filter(
            Q(day__contains=query4) |
            Q(faculty_name__contains=query4) |
            Q(nik__contains=query4) |
            Q(subject_name__contains=query4) |
            Q(event_name__contains=query4) |
            Q(section_name__contains=query4) |
            Q(room__contains=query4)
        )
        if (query2 == "" and query3 == "" and query4 == "" ):
            data = data1
        elif(query3 == "" and query4 == ""):
            data = data1 | data2
        elif(query4 == ""):
            data = data1 | data2 | data3
        else:
            data = data1 | data2 | data3 | data4

    else:
        data = models.NewKuliahRevise.objects.all()

    # now = datetime.now()
    # a = now.strftime("%w")
    # pengurang = int(a)-1
    # tanggal_mulai = now+timedelta(days=-pengurang)

    kuliah_senin = data.filter(day__contains="Monday").order_by("date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_senin:
        if iterate == 0:
            week1_tgl = item.date
            iterate = iterate +1
        else:
            if(item.date != week1_tgl):
                week2_tgl = item.date
                break;

    tanggal_senin= week1_tgl
    tanggal_senin2 = week2_tgl
    senin_1 = kuliah_senin.filter(date__date=tanggal_senin)
    senin_2 = kuliah_senin.filter(date__date=tanggal_senin2)
    kuliah_senin_slot1 = senin_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_senin_slot2 = senin_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_senin_slot3 = senin_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_senin_slot4 = senin_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_senin_slot5 = senin_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_senin2_slot1 = senin_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_senin2_slot2 = senin_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_senin2_slot3 = senin_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_senin2_slot4 = senin_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_senin2_slot5 = senin_2.filter(start_time__range=("17:50", "19:30"))

    kuliah_selasa = data.filter(day__contains="Tuesday").order_by("date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_selasa:
        if iterate == 0:
            week1_tgl = item.date
            iterate = iterate +1
        else:
            if(item.date != week1_tgl):
                week2_tgl = item.date
                break;
    tanggal_selasa =  week1_tgl
    tanggal_selasa2 = week2_tgl

    selasa_1 = kuliah_selasa.filter(date__date=tanggal_selasa)
    selasa_2 = kuliah_selasa.filter(date__date=tanggal_selasa2)
    kuliah_selasa_slot1 = selasa_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_selasa_slot2 = selasa_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_selasa_slot3 = selasa_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_selasa_slot4 = selasa_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_selasa_slot5 = selasa_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_selasa2_slot1 = selasa_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_selasa2_slot2 = selasa_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_selasa2_slot3 = selasa_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_selasa2_slot4 = selasa_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_selasa2_slot5 = selasa_2.filter(start_time__range=("17:50", "19:30"))

    kuliah_rabu = data.filter(day__contains="Wednesday").order_by("date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_rabu:
        if iterate == 0:
            week1_tgl = item.date
            iterate = iterate +1
        else:
            if(item.date != week1_tgl):
                # print("ee")
                week2_tgl = item.date
                break;
    tanggal_rabu =  week1_tgl
    tanggal_rabu2= week2_tgl
    rabu_1 = kuliah_rabu.filter(date__date=tanggal_rabu)
    rabu_2 = kuliah_rabu.filter(date__date=tanggal_rabu2)
    kuliah_rabu_slot1 = rabu_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_rabu_slot2 = rabu_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_rabu_slot3 = rabu_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_rabu_slot4 = rabu_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_rabu_slot5 = rabu_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_rabu2_slot1 = rabu_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_rabu2_slot2 = rabu_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_rabu2_slot3 = rabu_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_rabu2_slot4 = rabu_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_rabu2_slot5 = rabu_2.filter(start_time__range=("17:50", "19:30"))

    kuliah_kamis = data.filter(day__contains="Thursday").order_by("date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_kamis:
        if iterate == 0:
            week1_tgl = item.date
            iterate = iterate +1
        else:
            if(item.date != week1_tgl):
                # print("ee")
                week2_tgl = item.date
                break;
    tanggal_kamis =  week1_tgl
    tanggal_kamis2= week2_tgl
    kamis_1 = kuliah_kamis.filter(date__date=tanggal_kamis)
    kamis_2 = kuliah_kamis.filter(date__date=tanggal_kamis2)
    kuliah_kamis_slot1 = kamis_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_kamis_slot2 = kamis_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_kamis_slot3 = kamis_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_kamis_slot4 = kamis_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_kamis_slot5 = kamis_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_kamis2_slot1 = kamis_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_kamis2_slot2 = kamis_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_kamis2_slot3 = kamis_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_kamis2_slot4 = kamis_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_kamis2_slot5 = kamis_2.filter(start_time__range=("17:50", "19:30"))

    kuliah_jumat = data.filter(day__contains="Friday").order_by("date","start_time")
    week1_tgl = None
    week2_tgl = None
    iterate = 0
    for item in kuliah_jumat:
        if iterate == 0:
            week1_tgl = item.date
            iterate = iterate +1
        else:
            if(item.date != week1_tgl):
                # print("horeeee")
                week2_tgl = item.date
                break;
    tanggal_jumat =  week1_tgl
    tanggal_jumat2= week2_tgl
    jumat_1 = kuliah_jumat.filter(date__date=tanggal_jumat)
    jumat_2 = kuliah_jumat.filter(date__date=tanggal_jumat2)
    kuliah_jumat_slot1 = jumat_1.filter(start_time__range=("08:00", "10:59"))
    kuliah_jumat_slot2 = jumat_1.filter(start_time__range=("11:00", "12:59"))
    kuliah_jumat_slot3 = jumat_1.filter(start_time__range=("13:00", "15:59"))
    kuliah_jumat_slot4 = jumat_1.filter(start_time__range=("16:00", "17:49"))
    kuliah_jumat_slot5 = jumat_1.filter(start_time__range=("17:50", "19:30"))
    kuliah_jumat2_slot1 = jumat_2.filter(start_time__range=("08:00", "10:59"))
    kuliah_jumat2_slot2 = jumat_2.filter(start_time__range=("11:00", "12:59"))
    kuliah_jumat2_slot3 = jumat_2.filter(start_time__range=("13:00", "15:59"))
    kuliah_jumat2_slot4 = jumat_2.filter(start_time__range=("16:00", "17:49"))
    kuliah_jumat2_slot5 = jumat_2.filter(start_time__range=("17:50", "19:30"))

    #check empty slot for Senin
    if (tanggal_senin != None) and (tanggal_senin2 ==  None) :
        flag = False
        if tanggal_selasa != None:
            if tanggal_selasa < tanggal_senin:
                flag = True
        elif tanggal_rabu != None:
            if tanggal_rabu < tanggal_senin:
                flag = True
        elif tanggal_kamis != None:
            if tanggal_kamis < tanggal_senin:
                flag = True
        elif tanggal_jumat != None:
            if tanggal_jumat < tanggal_senin:
                flag = True

        if flag == True :
            tanggal_senin2 = tanggal_senin
            tanggal_senin = None
            kuliah_senin2_slot1 = kuliah_senin_slot1
            kuliah_senin2_slot2 = kuliah_senin_slot2
            kuliah_senin2_slot3 = kuliah_senin_slot3
            kuliah_senin2_slot4 = kuliah_senin_slot4
            kuliah_senin2_slot5 = kuliah_senin_slot5
            kuliah_senin_slot1 = []
            kuliah_senin_slot2 = []
            kuliah_senin_slot3 = []
            kuliah_senin_slot4 = []
            kuliah_senin_slot5 = []

    if (tanggal_selasa != None) and (tanggal_selasa2 ==  None) :
        flag = False
        if tanggal_rabu != None:
            if tanggal_rabu < tanggal_selasa:
                tanggal_selasa2 =  tanggal_selasa
        elif tanggal_kamis != None:
            if tanggal_kamis < tanggal_selasa:
                flag = True
        elif tanggal_jumat != None:
            if tanggal_jumat < tanggal_selasa:
                tanggal_selasa2 =  tanggal_selasa

        if flag == True :
            tanggal_selasa2 = tanggal_selasa
            tanggal_selasa = None
            kuliah_selasa2_slot1 = kuliah_selasa_slot1
            kuliah_selasa2_slot2 = kuliah_selasa_slot2
            kuliah_selasa2_slot3 = kuliah_selasa_slot3
            kuliah_selasa2_slot4 = kuliah_selasa_slot4
            kuliah_selasa2_slot5 = kuliah_selasa_slot5
            kuliah_selasa_slot1 = []
            kuliah_selasa_slot2 = []
            kuliah_selasa_slot3 = []
            kuliah_selasa_slot4 = []
            kuliah_selasa_slot5 = []

    if (tanggal_rabu != None) and (tanggal_rabu2 ==  None) :
        flag = False
        if tanggal_kamis != None:
            if tanggal_kamis < tanggal_rabu:
                flag = True
        elif tanggal_jumat != None:
            if tanggal_jumat < tanggal_rabu:
                flag = True

        if flag == True :
            tanggal_rabu2 = tanggal_rabu
            tanggal_rabu = None
            kuliah_rabu2_slot1 = kuliah_rabu_slot1
            kuliah_rabu2_slot2 = kuliah_rabu_slot2
            kuliah_rabu2_slot3 = kuliah_rabu_slot3
            kuliah_rabu2_slot4 = kuliah_rabu_slot4
            kuliah_rabu2_slot5 = kuliah_rabu_slot5
            kuliah_rabu_slot1 = []
            kuliah_rabu_slot2 = []
            kuliah_rabu_slot3 = []
            kuliah_rabu_slot4 = []
            kuliah_rabu_slot5 = []

    if (tanggal_kamis != None) and (tanggal_kamis2 ==  None) :
        flag = False
        if tanggal_jumat != None:
            if tanggal_jumat < tanggal_kamis:
                flag = True

        if flag == True :
            tanggal_kamis2 = tanggal_rabu
            tanggal_kamis = None
            kuliah_kamis2_slot1 = kuliah_kamis_slot1
            kuliah_kamis2_slot2 = kuliah_kamis_slot2
            kuliah_kamis2_slot3 = kuliah_kamis_slot3
            kuliah_kamis2_slot4 = kuliah_kamis_slot4
            kuliah_kamis2_slot5 = kuliah_kamis_slot5
            kuliah_kamis_slot1 = []
            kuliah_kamis_slot2 = []
            kuliah_kamis_slot3 = []
            kuliah_kamis_slot4 = []
            kuliah_kamis_slot5 = []

    if (tanggal_jumat != None) and (tanggal_jumat2 ==  None) :
        flag = False
        if tanggal_senin2!= None:
            if tanggal_senin2 < tanggal_jumat:
                flag = True
        elif tanggal_selasa2 != None:
            if tanggal_selasa2 < tanggal_jumat:
                flag = True
        elif tanggal_rabu2 != None:
            if tanggal_rabu2 < tanggal_jumat:
                flag = True
        elif tanggal_kamis2 != None:
            if tanggal_kamis2 < tanggal_jumat:
                flag = True

        if flag == True :
            tanggal_jumat2 = tanggal_jumat
            tanggal_jumat = None
            kuliah_jumat2_slot1 = kuliah_jumat_slot1
            kuliah_jumat2_slot2 = kuliah_jumat_slot2
            kuliah_jumat2_slot3 = kuliah_jumat_slot3
            kuliah_jumat2_slot4 = kuliah_jumat_slot4
            kuliah_jumat2_slot5 = kuliah_jumat_slot5
            kuliah_jumat_slot1 = []
            kuliah_jumat_slot2 = []
            kuliah_jumat_slot3 = []
            kuliah_jumat_slot4 = []
            kuliah_jumat_slot5 = []



    return render(request, 'kuliah.html', {
                "tanggal_senin":tanggal_senin,
                "tanggal_selasa":tanggal_selasa,
                "tanggal_rabu":tanggal_rabu,
                "tanggal_kamis":tanggal_kamis,
                "tanggal_jumat":tanggal_jumat,
                "tanggal_senin2":tanggal_senin2,
                "tanggal_selasa2":tanggal_selasa2,
                "tanggal_rabu2":tanggal_rabu2,
                "tanggal_kamis2":tanggal_kamis2,
                "tanggal_jumat2":tanggal_jumat2,

                "kuliah_senin_slot1":kuliah_senin_slot1,
                "kuliah_senin_slot2":kuliah_senin_slot2,
                "kuliah_senin_slot3":kuliah_senin_slot3,
                "kuliah_senin_slot4":kuliah_senin_slot4,
                "kuliah_senin_slot5":kuliah_senin_slot5,
                "kuliah_senin2_slot1":kuliah_senin2_slot1,
                "kuliah_senin2_slot2":kuliah_senin2_slot2,
                "kuliah_senin2_slot3":kuliah_senin2_slot3,
                "kuliah_senin2_slot4":kuliah_senin2_slot4,
                "kuliah_senin2_slot5":kuliah_senin2_slot5,

                "kuliah_selasa_slot1":kuliah_selasa_slot1,
                "kuliah_selasa_slot2":kuliah_selasa_slot2,
                "kuliah_selasa_slot3":kuliah_selasa_slot3,
                "kuliah_selasa_slot4":kuliah_selasa_slot4,
                "kuliah_selasa_slot5":kuliah_selasa_slot5,
                "kuliah_selasa2_slot1":kuliah_selasa2_slot1,
                "kuliah_selasa2_slot2":kuliah_selasa2_slot2,
                "kuliah_selasa2_slot3":kuliah_selasa2_slot3,
                "kuliah_selasa2_slot4":kuliah_selasa2_slot4,
                "kuliah_selasa2_slot5":kuliah_selasa2_slot5,

                "kuliah_rabu_slot1":kuliah_rabu_slot1,
                "kuliah_rabu_slot2":kuliah_rabu_slot2,
                "kuliah_rabu_slot3":kuliah_rabu_slot3,
                "kuliah_rabu_slot4":kuliah_rabu_slot4,
                "kuliah_rabu_slot5":kuliah_rabu_slot5,
                "kuliah_rabu2_slot1":kuliah_rabu2_slot1,
                "kuliah_rabu2_slot2":kuliah_rabu2_slot2,
                "kuliah_rabu2_slot3":kuliah_rabu2_slot3,
                "kuliah_rabu2_slot4":kuliah_rabu2_slot4,
                "kuliah_rabu2_slot5":kuliah_rabu2_slot5,

                "kuliah_kamis_slot1":kuliah_kamis_slot1,
                "kuliah_kamis_slot2":kuliah_kamis_slot2,
                "kuliah_kamis_slot3":kuliah_kamis_slot3,
                "kuliah_kamis_slot4":kuliah_kamis_slot4,
                "kuliah_kamis_slot5":kuliah_kamis_slot5,
                "kuliah_kamis2_slot1":kuliah_kamis2_slot1,
                "kuliah_kamis2_slot2":kuliah_kamis2_slot2,
                "kuliah_kamis2_slot3":kuliah_kamis2_slot3,
                "kuliah_kamis2_slot4":kuliah_kamis2_slot4,
                "kuliah_kamis2_slot5":kuliah_kamis2_slot5,

                "kuliah_jumat_slot1":kuliah_jumat_slot1,
                "kuliah_jumat_slot2":kuliah_jumat_slot2,
                "kuliah_jumat_slot3":kuliah_jumat_slot3,
                "kuliah_jumat_slot4":kuliah_jumat_slot4,
                "kuliah_jumat_slot5":kuliah_jumat_slot5,
                "kuliah_jumat2_slot1":kuliah_jumat2_slot1,
                "kuliah_jumat2_slot2":kuliah_jumat2_slot2,
                "kuliah_jumat2_slot3":kuliah_jumat2_slot3,
                "kuliah_jumat2_slot4":kuliah_jumat2_slot4,
                "kuliah_jumat2_slot5":kuliah_jumat2_slot5,
        }
    )



def uploadsap(request):
    if "GET" == request.method:
        return render(request, 'uploadjadwal.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row

        counter = 0
        counter_valid = 0
        models.Kuliah.objects.all().delete()
        for row in worksheet.iter_rows():
            #this is in one row

            row_data = list()
            for cell in row:
                # print(cell)
                row_data.append(str(cell.value))

            excel_data.append(row_data)
            # Format data :
            # Class Date | Day | Start Time | End Time | Full Name | Initial | Subject Code | Subject Name | Cohort Name | Room No |

            #TODO masukin ke DB row_data
            if counter >0 :
                #filter initial name lecturer that having class in STEM
                init = row_data[4]

                if (init == "MYG" or init == "JAN" or init == "DKN" or init == "MZS" or init == "SPR" or init == "VAS" or init == "ERW" or init == "AAL" or init == "JIL" or init == "STW"or
                    init == "SAM" or init == "EAR" or init == "LJD" or init == "BYP" or init == "INB" or init == "FSH" or init == "YSA" or init == "YAF" or init == "HAF" or init == "ILS" or
                    init == "NYT" or init == "RTK" or init == "FSC" or init == "ZSS" or init == "NMA" or init == "PPU" or init == "SKT" or init == "AAN" or init == "AAW" or init == "HLW" or
                    init == "PNM" or init == "ADU" or init == "IHP" or init == "SMK" or init == "FZS" or init == "RIY" or init == "MLT" or init == "AFN" or init == "SDR" or init == "CM" or
                    init == "GAA" or init == "STV" or init == "ARKO" or init == "TJP" or init == "NAU" or init == "YAD" or init == "PLT" or init == "FTH" or init == "ATE" or init == "JIK" or
                    init == "AUL" or init == "FS" ) :

                    ## TODO gnti cohort name dengan filter cohort_name, initial, subject_code
                    temp_cohort = row_data[7]
                    if(temp_cohort == "S1 BUSMATH Angkt 2017 Sem 5"):
                        temp_cohort = "BM 2017A"
                    elif temp_cohort == "S1 BUSMATH Angkt 2018 Sem 3_A":
                        temp_cohort = "BM 2018A"
                    elif (temp_cohort == "S1 BUSMATH Angkt 2018 Sem 3_B"):
                        temp_cohort = "BM 2018B"
                    elif (temp_cohort == "S1 BUSMATH Angkt 2018 Sem 3_C"):
                        temp_cohort = "BM 2018C"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_A" and row_data[5] == "MBSU1101" and row_data[4]=="MYG"):
                        temp_cohort = "BM 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_A" and row_data[5] == "MBSU1125" and row_data[4]=="AFR"):
                        temp_cohort = "BM 2019A/ESE 2018A"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_A"):
                        temp_cohort = "BM 2019A"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_B"):
                        temp_cohort = "BM 2019B"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_C"):
                        temp_cohort = "BM 2019C"
                    elif (temp_cohort == "S1 CSE Angkt 2017 Sem 5"):
                        temp_cohort = "CSE 2017A"
                    elif (temp_cohort == "S1 CSE Angkt 2018 Sem 3" and row_data[5] == "REEU2107" and row_data[4]=="AAN"):
                        temp_cohort = "REE 2018A/PDE 2018A/CSE 2018A"
                    elif (temp_cohort == "S1 CSE Angkt 2018 Sem 3" and row_data[5] == "UNIU1W01" and row_data[4]=="NAU"):
                        temp_cohort = "CSE 2018A/ESE 2018A/REE 2018A"
                    elif (temp_cohort == "S1 CSE Angkt 2018 Sem 3"):
                        temp_cohort = "CSE 2018A"
                    elif (temp_cohort == "S1 CSE angkt 2019 sem 1" and row_data[5] == "PDEU1151" and row_data[4]=="ADU"):
                        temp_cohort = "PDE 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 CSE angkt 2019 sem 1"):
                        temp_cohort = "CSE 2019A"
                    elif (temp_cohort == "S1 FBT angkt 2019 sem 1" and row_data[5] == "STEM1102" and row_data[4]=="SDR"):
                        temp_cohort = "FBT 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 FBT angkt 2019 sem 1"):
                        temp_cohort = "FBT 2019A"
                    elif (temp_cohort == "S1 FBT-LAB1 Angkt 2017 Sem 5"):
                        temp_cohort = "FBT 2017A"
                    elif (temp_cohort == "S1 FBT-LAB1 Angkt 2018 Sem 3" and row_data[5] == "FBTE2116" and row_data[4]=="MZS"):
                        temp_cohort = "FBT 2018A/CSE 2017A"
                    elif (temp_cohort == "S1 FBT-LAB1 Angkt 2018 Sem 3"):
                        temp_cohort = "FBT 2018A"
                    elif (temp_cohort == "S1 FBT-LAB2 Angkt 2017 Sem 5"):
                        temp_cohort = "FBT 2017A"
                    elif (temp_cohort == "S1 FBT-LAB2 Angkt 2018 Sem 3"):
                        temp_cohort = "FBT 2018A"
                    elif (temp_cohort == "S1 PDE Angkt 2018 Sem 3"):
                        temp_cohort = "PDE 2018A"
                    elif (temp_cohort == "S1 PDE angkt 2019 sem 1"):
                        temp_cohort = "PDE 2019A"
                    elif (temp_cohort == "S1 REE Angkatan 2018 Sem 3" and row_data[5] == "CSEE2356" and row_data[4]=="ERW"):
                        temp_cohort = "REE 2018A/PDE 2018A"
                    elif (temp_cohort == "S1 REE Angkatan 2018 Sem 3" and row_data[5] == "REEU2109" and row_data[4]=="LJD"):
                        temp_cohort = "REE 2018A/PDE 2018A"
                    elif (temp_cohort == "S1 REE Angkatan 2018 Sem 3"):
                        temp_cohort = "REE 2018A"
                    elif (temp_cohort == "S1 REE Angkt 2017 Sem 5" and row_data[5] == "STEM3109" and row_data[4]=="YSA"):
                        temp_cohort = "BM 2017A/FBT 2017A/REE 2017A/CSE 2017A/ESE 2017A"
                    elif (temp_cohort == "S1 REE Angkt 2017 Sem 5" and row_data[5] == "STEM3109" and row_data[4]=="FS"):
                        temp_cohort = "BM 2017A/FBT 2017A/REE 2017A/CSE 2017A/ESE 2017A"
                    elif (temp_cohort == "S1 REE Angkt 2017 Sem 5"):
                        temp_cohort = "REE 2017A"
                    elif (temp_cohort == "S1 RENEWABLE ENERGY Angkatan 2019 Sem 1" and row_data[5] == "UNIU1W05" and row_data[4]=="AAW"):
                        temp_cohort = "REE 2019A/PDE 2019A"
                    elif (temp_cohort == "S1 RENEWABLE ENERGY Angkatan 2019 Sem 1" and row_data[5] == "REEU1101" and row_data[4]=="NMA"):
                        temp_cohort = "REE 2019A/PDE 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 RENEWABLE ENERGY Angkatan 2019 Sem 1" and row_data[5] == "PDEA1101" and row_data[4]=="ZSS"):
                        temp_cohort = "REE 2019A/PDE 2019A"
                    elif (temp_cohort == "S1 RENEWABLE ENERGY Angkatan 2019 Sem 1"):
                        temp_cohort = "REE 2019A"
                    elif (temp_cohort == "S1 SE Angkt 2017 Sem 5"):
                        temp_cohort = "ESE 2017A"
                    elif (temp_cohort == "S1 SE Angkt 2018 Sem 3"):
                        temp_cohort = "ESE 2018A"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and row_data[5] == "REEU1103" and row_data[4]=="YSA"):
                        temp_cohort = "REE 2019A/ESE 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and row_data[5] == "REEU1103" and row_data[4]=="FSH"):
                        temp_cohort = "REE 2019A/ESE 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and row_data[5] == "UNIU1W05" and row_data[4]=="SKT"):
                        temp_cohort = "CSE 2019/ESE 2019"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and row_data[5] == "STEM1102" and row_data[4]=="SDR"):
                        temp_cohort = "CSE 2019/ESE 2019"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and row_data[5] == "MBSU1101" and row_data[4]=="DKN"):
                        temp_cohort = "REE 2019A/PDE 2019A/ESE 2019A"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1"):
                        temp_cohort = "ESE 2019A"

                    item_kuliah = models.Kuliah(
                        class_date=row_data[0],
                        start_time=row_data[1],
                        end_time=row_data[2],
                        full_name=row_data[3],
                        initial=row_data[4],
                        subject_code=row_data[5],
                        subject_name=row_data[6],
                        cohort_name=temp_cohort,
                        room_no=row_data[9],
                        day=datetime.strptime(row_data[0], '%Y-%m-%d %H:%M:%S').strftime("%A"),
                    )
                    resp = item_kuliah.save()
                    counter_valid = counter_valid + 1
            #     print(resp)


            # print(row_data)
            counter = counter + 1

        return render(request, 'uploadjadwal.html', {
                "excel_data":excel_data,
                "length_data":counter_valid
            }
        )

def newuploadsap(request):
    if "GET" == request.method:
        return render(request, 'uploadjadwal.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row

        counter = 0
        counter_valid = 0

        models.NewKuliahRevise.objects.all().delete()

        # Lecturer that having class in STEM
        nik_list = ["240", "367", "441", "526", "572", "608", "613", "620", "667", "680", "684", "687", "688", "699", "700", "705", "711", "725", "735", "736", "741", "755", "758", "761", "765", "783", "797", "806", "822", "835", "858", "871", "897", "914", "942", "946", "969", "970", "972", "983", "7001", "7025", "7058", "7129", "7130"]

        for row in worksheet.iter_rows():
            #this is in one row
            row_data = list()
            for cell in row:
                # print(cell)
                row_data.append(str(cell.value))

            excel_data.append(row_data)
            # Format data :
            # Class Date | Day | Start Time | End Time | Full Name | Initial | Subject Code | Subject Name | Cohort Name | Room No |

            # New Format Data:
            # Program ID | Program Name | Section ID | Section Name | Day | Date | Subject Name | Event Obj | Event Name | Start Time | End Time | Capacitye | Location | N I K | Faculty Name

            #TODO masukin ke DB row_data
            if counter > 0 :
                #filter initial name lecturer that having class in STEM
                nik = row_data[13]
                # print("nik", nik)
                if nik in nik_list:
                    # print("nik in", nik)
                    ## TODO gnti cohort name dengan filter cohort_name, initial, subject_code
                    temp_cohort = row_data[3]
                    if(temp_cohort == "S1 BUSMATH Angkt 2017 Sem 5"):
                        temp_cohort = "BM 2017A"
                    elif temp_cohort == "S1 BUSMATH Angkt 2018 Sem 3_A":
                        temp_cohort = "BM 2018A"
                    elif (temp_cohort == "S1 BUSMATH Angkt 2018 Sem 3_B"):
                        temp_cohort = "BM 2018B"
                    elif (temp_cohort == "S1 BUSMATH Angkt 2018 Sem 3_C"):
                        temp_cohort = "BM 2018C"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_A" and "MBSU1101" in row_data[6] and nik == "700"):
                        temp_cohort = "BM 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_A" and "MBSU1125" in row_data[6] and nik == "942"):
                        temp_cohort = "BM 2019A/ESE 2018A"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_A"):
                        temp_cohort = "BM 2019A"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_B"):
                        temp_cohort = "BM 2019B"
                    elif (temp_cohort == "S1 BUSMATH angkt 2019 sem 1_C"):
                        temp_cohort = "BM 2019C"
                    elif (temp_cohort == "S1 CSE Angkt 2017 Sem 5"):
                        temp_cohort = "CSE 2017A"
                    elif (temp_cohort == "S1 CSE Angkt 2018 Sem 3" and "REEU2107" in row_data[6] and nik == "797"):
                        temp_cohort = "REE 2018A/PDE 2018A/CSE 2018A"
                    elif (temp_cohort == "S1 CSE Angkt 2018 Sem 3" and "UNIU1W01" in row_data[6] and nik == "7058"):
                        temp_cohort = "CSE 2018A/ESE 2018A/REE 2018A"
                    elif (temp_cohort == "S1 CSE Angkt 2018 Sem 3"):
                        temp_cohort = "CSE 2018A"
                    elif (temp_cohort == "S1 CSE angkt 2019 sem 1" and "PDEU1151" in row_data[6] and nik == "871"):
                        temp_cohort = "PDE 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 CSE angkt 2019 sem 1"):
                        temp_cohort = "CSE 2019A"
                    elif (temp_cohort == "S1 FBT angkt 2019 sem 1" and "STEM1102" in row_data[6] and nik == "7001"):
                        temp_cohort = "FBT 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 FBT angkt 2019 sem 1"):
                        temp_cohort = "FBT 2019A"
                    elif (temp_cohort == "S1 FBT-LAB1 Angkt 2017 Sem 5"):
                        temp_cohort = "FBT 2017A"
                    elif (temp_cohort == "S1 FBT-LAB1 Angkt 2018 Sem 3" and "FBTE2116" in row_data[6] and nik == "699"):
                        temp_cohort = "FBT 2018A/CSE 2017A"
                    elif (temp_cohort == "S1 FBT-LAB1 Angkt 2018 Sem 3"):
                        temp_cohort = "FBT 2018A"
                    elif (temp_cohort == "S1 FBT-LAB2 Angkt 2017 Sem 5"):
                        temp_cohort = "FBT 2017A"
                    elif (temp_cohort == "S1 FBT-LAB2 Angkt 2018 Sem 3"):
                        temp_cohort = "FBT 2018A"
                    elif (temp_cohort == "S1 PDE Angkt 2018 Sem 3"):
                        temp_cohort = "PDE 2018A"
                    elif (temp_cohort == "S1 PDE angkt 2019 sem 1"):
                        temp_cohort = "PDE 2019A"
                    elif (temp_cohort == "S1 REE Angkatan 2018 Sem 3" and "CSEE2356" in row_data[6] and nik == "705"):
                        temp_cohort = "REE 2018A/PDE 2018A"
                    elif (temp_cohort == "S1 REE Angkatan 2018 Sem 3" and "REEU2109" in row_data[6] and nik == "620"):
                        temp_cohort = "REE 2018A/PDE 2018A"
                    elif (temp_cohort == "S1 REE Angkatan 2018 Sem 3"):
                        temp_cohort = "REE 2018A"
                    elif (temp_cohort == "S1 REE Angkt 2017 Sem 5" and "STEM3109" in row_data[6] and nik == "688"):
                        temp_cohort = "BM 2017A/FBT 2017A/REE 2017A/CSE 2017A/ESE 2017A"
                    elif (temp_cohort == "S1 REE Angkt 2017 Sem 5" and "STEM3109" in row_data[6] and nik == "761"):
                        temp_cohort = "BM 2017A/FBT 2017A/REE 2017A/CSE 2017A/ESE 2017A"
                    elif (temp_cohort == "S1 REE Angkt 2017 Sem 5"):
                        temp_cohort = "REE 2017A"
                    elif (temp_cohort == "S1 RENEWABLE ENERGY Angkatan 2019 Sem 1" and "UNIU1W05" in row_data[6] and nik == "806"):
                        temp_cohort = "REE 2019A/PDE 2019A"
                    elif (temp_cohort == "S1 RENEWABLE ENERGY Angkatan 2019 Sem 1" and "REEU1101" in row_data[6] and nik == "765"):
                        temp_cohort = "REE 2019A/PDE 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 RENEWABLE ENERGY Angkatan 2019 Sem 1" and "PDEA1101" in row_data[6] and nik == "711"):
                        temp_cohort = "REE 2019A/PDE 2019A"
                    elif (temp_cohort == "S1 RENEWABLE ENERGY Angkatan 2019 Sem 1"):
                        temp_cohort = "REE 2019A"
                    elif (temp_cohort == "S1 SE Angkt 2017 Sem 5"):
                        temp_cohort = "ESE 2017A"
                    elif (temp_cohort == "S1 SE Angkt 2018 Sem 3"):
                        temp_cohort = "ESE 2018A"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and "REEU1103" in row_data[6] and nik == "688"):
                        temp_cohort = "REE 2019A/ESE 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and "REEU1103" in row_data[6] and nik == "761"):
                        temp_cohort = "REE 2019A/ESE 2019A/CSE 2019A"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and "UNIU1W05" in row_data[6] and nik == "783"):
                        temp_cohort = "CSE 2019/ESE 2019"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and "STEM1102" in row_data[6] and nik == "7001"):
                        temp_cohort = "CSE 2019/ESE 2019"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1" and "MBSU1101" in row_data[6] and nik == "687"):
                        temp_cohort = "REE 2019A/PDE 2019A/ESE 2019A"
                    elif (temp_cohort == "S1 SE angkt 2019 sem 1"):
                        temp_cohort = "ESE 2019A"

                    item_kuliah = models.NewKuliahRevise (
                        program_id      = row_data[0],
                        program_name    = row_data[1],
                        section_id      = row_data[2],
                        section_name    = temp_cohort,
                        day             = row_data[4],
                        date            = row_data[5],
                        subject_name    = row_data[6],
                        event_obj       = row_data[7],
                        event_name      = row_data[8],
                        start_time      = row_data[9],
                        end_time        = row_data[10],
                        capacity        = row_data[11],
                        location        = row_data[12],
                        nik             = row_data[13],
                        faculty_name    = row_data[14],
                        room            = row_data[15]
                    )
                    resp = item_kuliah.save()
                    counter_valid = counter_valid + 1
                # print(resp)


            # print(row_data)
            counter = counter + 1

        return render(request, 'uploadjadwal.html', {
                "excel_data":excel_data,
                "length_data":counter_valid
            }
        )


# Create your views here.
class JadwalCreate(CreateView):  # new
    model = models.JadwalHarian
    form_class = forms.JadwalForm
    template_name = 'jadwalcreate.html'

    def get_context_data(self, **kwargs):
        context = super(JadwalCreate, self).get_context_data(**kwargs)
        context["title"] = "Penambahan Jadwal Kuliah STEM Prasetiya Mulya"
        return context

class RoomViewSet(viewsets.ModelViewSet):
    """Handles CRUD for model Room"""

    serializer_class = serializers.RoomSerializer
    queryset = models.Room.objects.all()

class LecturerViewSet(viewsets.ModelViewSet):
    """Handles CRUD for model Lecturer"""

    serializer_class = serializers.LecturerSerializer
    queryset = models.Lecturer.objects.all()

class StudentViewSet(viewsets.ModelViewSet):
    """Handles CRUD for model Student"""

    serializer_class = serializers.StudentSerializer
    queryset = models.Student.objects.all()

class ScheduleViewSet(viewsets.ModelViewSet):
    """Handles CRUD for model Schedule"""

    serializer_class = serializers.ScheduleSerializer
    queryset = models.RoomSlot.objects.all()

class KuliahViewSet(viewsets.ModelViewSet):
    """Handles CRUD for model Schedule"""

    serializer_class = serializers.KuliahSerializer
    queryset = models.Kuliah.objects.all()

class JadwalViewSet(viewsets.ModelViewSet):
    """Handles CRUD for model Schedule"""

    serializer_class = serializers.JadwalSerializer
    queryset = models.JadwalHarian.objects.all()

class StemView(TemplateView):
    template_name = 'stem.html'

class CalendarView(TemplateView):
    template_name = 'calendar.html'

class CalendarNewView(TemplateView):
    template_name = 'calendarnew.html'
